import json
from pathlib import Path

from openpyxl import load_workbook

HERE = Path(__file__).resolve().parent
XLSX = HERE / "saransk_flats.xlsx"
OUT = HERE / "housing_dataset.json"
COLS = ["РАЙОН", "КОМНАТЫ", "ПЛОЩАДЬ", "ЭТАЖ", "АДРЕС", "ЦЕНА", "ССЫЛКА"]


def main() -> None:
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else "" for h in next(rows)]
    idx = {c: header.index(c) for c in COLS}
    data = []
    for row in rows:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        data.append({c: row[idx[c]] if idx[c] < len(row) else None for c in COLS})
    wb.close()
    OUT.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"OK {len(data)} записей → {OUT}")


if __name__ == "__main__":
    main()
