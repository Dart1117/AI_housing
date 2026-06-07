import asyncio
import random
import re
import json
from collections import namedtuple
from urllib.parse import urlsplit, urlunsplit

from playwright.async_api import async_playwright, TimeoutError as PwTimeout
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


SEARCHES = [
    {"district": "Пролетарский", "url": "https://www.avito.ru/saransk/kvartiry/prodam-ASgBAgICAUSSA8YQ?district=54054&s=104"},
    {"district": "Октябрьский", "url": "https://www.avito.ru/saransk/kvartiry/prodam-ASgBAgICAUSSA8YQ?district=54055&s=104"},
    {"district": "Ленинский", "url": "https://www.avito.ru/saransk/kvartiry/prodam-ASgBAgICAUSSA8YQ?district=54056&s=104"},
]

OUTPUT_FILE = "saransk_flats.xlsx"
MAX_PAGES = 60
PAGE_DELAY = (3, 6)
DETAIL_TIMEOUT_MS = 25000
DETAIL_DELAY = (0.7, 1.6)

# Маркер кэша: страница объявления недоступна (IP / антибот) — не дёргать повторно в одном прогоне
_DETAIL_BLOCKED = "__DETAIL_BLOCKED__"

STEALTH_JS = """
() => {
    Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
    Object.defineProperty(navigator, 'plugins', {
        get: () => [1, 2, 3, 4, 5]
    });
    Object.defineProperty(navigator, 'languages', {
        get: () => ['ru-RU', 'ru', 'en-US', 'en']
    });
    window.chrome = { runtime: {} };
    const orig = window.navigator.permissions.query;
    window.navigator.permissions.query = (parameters) =>
        parameters.name === 'notifications'
            ? Promise.resolve({ state: Notification.permission })
            : orig(parameters);
}
"""

USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
]


def norm_spaces(text: str) -> str:
    if not text:
        return ""
    t = str(text).replace("\u00a0", " ").replace("\u2002", " ").replace("\u202f", " ")
    return re.sub(r"\s+", " ", t).strip()


def norm_lower(text: str) -> str:
    return norm_spaces(text).lower().replace("ё", "е")


# Не писать в колонку «РАЙОН»: город, регион, страна (это не район).
FORBIDDEN_DISTRICT_LABELS = frozenset(
    {
        "саранск",
        "россия",
        "рф",
        "мордовия",
        "республика мордовия",
        "республика",
    }
)


def is_forbidden_district_label(label: str) -> bool:
    t = norm_lower(label)
    if not t:
        return True
    if t in FORBIDDEN_DISTRICT_LABELS:
        return True
    if "мордовия" in t and len(t) < 25:
        return True
    if t.startswith("республика ") and "мордовия" in t:
        return True
    return False


def sanitize_district(label: str) -> str:
    s = norm_spaces(label)
    if not s:
        return ""
    if is_forbidden_district_label(s):
        return ""
    return s


CANONICAL_DISTRICTS = {
    "пролетарский": "Пролетарский",
    "октябрьский": "Октябрьский",
    "ленинский": "Ленинский",
}

DISTRICT_ALIASES = {
    **CANONICAL_DISTRICTS,
    "луховка": "Луховка",
    "химмаш": "Химмаш",
    "светотехстрой": "Светотехстрой",
    "юго-запад": "Юго-Запад",
    "югозапад": "Юго-Запад",
    "юго запад": "Юго-Запад",
    "центр": "Центр",
    "посоп": "Посоп",
    "ялга": "Ялга",
    "николаевка": "Николаевка",
    "горяйновка": "Горяйновка",
    "озерный": "Озерный",
    "озерный пос": "Озерный",
    "озерный поселок": "Озерный",
    "п озерный": "Озерный",
    "пос озерный": "Озерный",
    "поселок озерный": "Озерный",
}

# Длинные ключи раньше (чтобы «юго запад» победил «юго»).
_ALIAS_KEYS_SORTED = sorted(DISTRICT_ALIASES.keys(), key=len, reverse=True)


# Парсинг карточки
def parse_rooms(title: str) -> str:
    t = title.lower()
    if "студия" in t:
        return "Студия"
    m = re.search(r"\b(\d+)\s*[-]?\s*к\b", t, re.I)
    if m:
        return m.group(1)
    m = re.search(r"(\d+)[- ]?комн", t)
    return m.group(1) if m else ""


def parse_area(text: str) -> str:
    m = re.search(r"(\d+(?:[.,]\d+)?)\s*м²", text)
    if not m:
        m = re.search(r"(\d+(?:[.,]\d+)?)\s*кв\.?\s*м", text, re.I)
    return m.group(1).replace(",", ".") if m else ""


def parse_floor(text: str) -> str:
    m = re.search(r"(\d+)\s*/\s*(\d+)\s*эт", text, re.I)
    if m:
        return f"{m.group(1)}/{m.group(2)}"
    m = re.search(r"(\d+)\s*этаж", text, re.I)
    return m.group(1) if m else ""


def parse_price(text: str) -> str:
    if not text:
        return ""
    m = re.search(r"([\d\s\u00a0]+)\s*₽", text)
    if m:
        return re.sub(r"[\s\u00a0]", "", m.group(1))
    m = re.search(r"([\d\s\u00a0]{5,})\s*(?:руб|₽)", text, re.I)
    if m:
        return re.sub(r"[\s\u00a0]", "", m.group(1))
    nums = re.findall(r"\d[\d\s\u00a0]{4,}", text)
    if nums:
        best = max(nums, key=lambda s: len(re.sub(r"[\s\u00a0]", "", s)))
        return re.sub(r"[\s\u00a0]", "", best)
    return ""


def normalize_avito_link(link: str) -> str:
    if not link:
        return ""
    try:
        p = urlsplit(link)
        return urlunsplit((p.scheme, p.netloc, p.path, "", ""))
    except Exception:
        return link.strip()


def extract_district_hint(text: str) -> str:
    t = norm_spaces(text)
    if not t:
        return ""
    t_norm = norm_lower(t)

    m = re.search(
        r"\b(?:р-?н\.?|район|округ|мкр\.?|мкр|микрорайон)\s+([а-яё\-]+(?:\s+[а-яё\-]+)?)\b",
        t_norm,
        re.I,
    )
    if m:
        return m.group(1).strip().lower().replace("ё", "е")

    for key in _ALIAS_KEYS_SORTED:
        if re.search(rf"(?<![а-яё0-9]){re.escape(key)}(?![а-яё0-9])", t_norm, re.I):
            return key
    return ""


def hint_to_display(hint_key: str) -> str:
    if not hint_key:
        return ""
    hk = hint_key.strip().lower().replace("ё", "е")
    if hk in DISTRICT_ALIASES:
        return DISTRICT_ALIASES[hk]
    return " ".join(w.capitalize() for w in re.split(r"\s+", hk) if w)


def infer_microdistrict_from_street(blob: str) -> str:
    t = norm_lower(blob)
    if re.search(r"косарева|северо[- ]?восточн", t):
        return "Химмаш"
    return ""


def resolve_district(address: str, desc: str, title: str, evidence: str, search_district: str) -> str:
    blob = norm_spaces(f"{address} {desc} {title} {evidence}")
    hint = extract_district_hint(blob)
    if hint:
        out = hint_to_display(hint)
        return sanitize_district(out)

    micro = infer_microdistrict_from_street(blob)
    if micro:
        return micro

    sd = sanitize_district(search_district)
    return sd


def flatten_address_cell(text: str) -> str:
    if not text:
        return ""
    t = str(text).replace("\u00a0", " ").replace("\u2002", " ").replace("\u202f", " ")
    t = re.sub(r"\s*\n+\s*", " | ", t.strip())
    return norm_spaces(t)


def looks_like_street_address(text: str) -> bool:
    t = norm_lower(text)
    if not t:
        return False
    return bool(
        re.search(
            r"\b(ул\.?|улица|пр-?т\.?|просп(ект)?|ш\.?|шоссе|б-р|бульвар|пер\.?|переулок|пл\.?|площадь|наб\.?|набережная|проезд|рп\.?|пос\.?|пгт\.?|дом|д\.|стр\.)\b",
            t,
            re.I,
        )
    )


def looks_jk_only(address: str) -> bool:
    t = norm_lower(address)
    if "жк" not in t:
        return False
    return not looks_like_street_address(address)


def extract_address_line_from_text(body: str) -> str:
    t = (body or "").replace("\u00a0", " ").replace("\u2002", " ").replace("\u202f", " ")
    for line in t.splitlines():
        ln = line.strip()
        if len(ln) < 8:
            continue
        lnn = norm_lower(ln)
        if "саранск" in lnn and looks_like_street_address(lnn):
            return ln
    for line in t.splitlines():
        ln = line.strip()
        if len(ln) < 8:
            continue
        if looks_like_street_address(ln):
            return ln
    m = re.search(r"\bСаранск\b[^\n]{0,140}", t, re.I)
    return m.group(0).strip() if m else ""


def page_looks_blocked(body: str) -> bool:
    b = norm_lower(body)
    return any(
        x in b
        for x in (
            "доступ ограничен",
            "проблема с ip",
            "проблема с ip-",
            "доступ временно ограничен",
            "подтвердите, что вы не робот",
            "captcha",
            "капча",
        )
    )


def pick_best_address(candidates: list[str]) -> str:
    cands = [norm_spaces(c) for c in candidates if c and len(norm_spaces(c)) >= 4]
    if not cands:
        return ""
    street = [c for c in cands if looks_like_street_address(c)]
    pool = street if street else cands
    return max(pool, key=len)


DetailParseResult = namedtuple("DetailParseResult", "address evidence blocked")


async def parse_detail_page(detail_page, link: str) -> DetailParseResult:
    if not link:
        return DetailParseResult("", "", False)

    try:
        await detail_page.goto(link, wait_until="domcontentloaded", timeout=DETAIL_TIMEOUT_MS)
        await asyncio.sleep(random.uniform(*DETAIL_DELAY))
    except Exception:
        return DetailParseResult("", "", False)

    try:
        body = await detail_page.inner_text("body")
    except Exception:
        return DetailParseResult("", "", False)

    if page_looks_blocked(body):
        return DetailParseResult("", "", True)

    candidates: list[str] = []
    selectors = [
        '[data-marker="item-view/location"]',
        '[data-marker="item-address"]',
        '[itemprop="address"]',
        '[data-marker="item-view/item-address"]',
        '[data-marker="item-view/item-address-string"]',
        'div[class*="style-item-address"]',
        'div[class*="item-address"]',
    ]
    for sel in selectors:
        try:
            try:
                await detail_page.wait_for_selector(sel, timeout=2500)
            except Exception:
                pass
            els = await detail_page.query_selector_all(sel)
            for el in els[:5]:
                try:
                    txt = norm_spaces(await el.inner_text())
                    if len(txt) >= 6:
                        candidates.append(txt)
                except Exception:
                    continue
        except Exception:
            continue

    address = pick_best_address(candidates)

    if not looks_like_street_address(address):
        try:
            scripts = await detail_page.query_selector_all('script[type="application/ld+json"]')
            for s in scripts:
                raw = (await s.inner_text()).strip()
                if not raw:
                    continue
                try:
                    data = json.loads(raw)
                except Exception:
                    continue

                def walk(x):
                    if isinstance(x, dict):
                        yield x
                        for v in x.values():
                            yield from walk(v)
                    elif isinstance(x, list):
                        for i in x:
                            yield from walk(i)

                for obj in walk(data):
                    addr = obj.get("address") if isinstance(obj, dict) else None
                    if isinstance(addr, dict):
                        parts = []
                        for k in ("streetAddress", "addressLocality", "addressRegion", "postalCode"):
                            v = addr.get(k)
                            if v and isinstance(v, str):
                                parts.append(v.strip())
                        cand = ", ".join(p for p in parts if p)
                        if cand:
                            candidates.append(cand)
                    elif isinstance(addr, str) and addr.strip():
                        candidates.append(addr.strip())
        except Exception:
            pass
        address = pick_best_address(candidates + ([address] if address else []))

    evidence = extract_address_line_from_text(body)
    if not evidence:
        evidence = body[:4000]

    if (not looks_like_street_address(address)) and looks_like_street_address(evidence):
        address = evidence

    return DetailParseResult(address, evidence, False)


async def human_scroll(page):
    for _ in range(random.randint(3, 6)):
        await page.mouse.wheel(0, random.randint(300, 700))
        await asyncio.sleep(random.uniform(0.2, 0.6))


async def scroll_list_until_stable(page, max_rounds: int = 35) -> None:
    last_n = -1
    stable = 0
    for _ in range(max_rounds):
        items = await page.query_selector_all('[data-marker="item"]')
        n = len(items)
        if n <= 0:
            await asyncio.sleep(0.35)
            continue
        if n == last_n:
            stable += 1
            if stable >= 4:
                break
        else:
            stable = 0
            last_n = n
        await page.mouse.wheel(0, random.randint(1400, 2200))
        await asyncio.sleep(random.uniform(0.25, 0.55))


async def parse_listing_page(
    page, district: str, detail_page=None, detail_cache: dict | None = None
) -> list[dict]:
    records = []

    await human_scroll(page)
    await scroll_list_until_stable(page)
    await asyncio.sleep(random.uniform(0.35, 0.9))

    items = await page.query_selector_all('[data-marker="item"]')
    if not items:
        items = await page.query_selector_all('div[class*="iva-item-root"]')

    for item in items:
        try:
            title_el = await item.query_selector('[itemprop="name"], [data-marker="item-title"]')
            title = (await title_el.inner_text()).strip() if title_el else ""

            link_el = await item.query_selector('a[data-marker="item-title"]')
            if not link_el:
                link_el = await item.query_selector('a[href*="/saransk/"]')
            href = await link_el.get_attribute("href") if link_el else ""
            link = f"https://www.avito.ru{href}" if href and href.startswith("/") else href
            link = normalize_avito_link(link)
            if "/kvartiry/" not in link:
                continue

            price_el = await item.query_selector(
                '[data-marker="item-price"], [itemprop="price"], meta[itemprop="price"]'
            )
            price_text = ""
            if price_el:
                price_text = (
                    await price_el.inner_text()
                    if await price_el.get_attribute("content") is None
                    else await price_el.get_attribute("content") or ""
                )
            price = parse_price(price_text) or re.sub(r"[^\d]", "", price_text)

            desc_el = await item.query_selector(
                '[data-marker="item-specific-params"], [class*="geo-address"], [class*="params"]'
            )
            desc = (await desc_el.inner_text()).strip() if desc_el else ""

            full_text = f"{title} {desc}"

            addr_parts: list[str] = []
            for sel in (
                '[data-marker="item-address"]',
                '[data-marker="geo-address"]',
                '[class*="geo-address"]',
            ):
                try:
                    ae = await item.query_selector(sel)
                    if ae:
                        txt = norm_spaces(await ae.inner_text())
                        if txt:
                            addr_parts.append(txt)
                except Exception:
                    continue
            address = pick_best_address(addr_parts)

            evidence_extra = ""
            row_district = resolve_district(address, desc, title, evidence_extra, district)

            need_detail = (not address) or looks_jk_only(address) or (not row_district)

            if need_detail and detail_page and detail_cache is not None and link:
                cached = detail_cache.get(link)
                if cached == _DETAIL_BLOCKED:
                    pass
                elif isinstance(cached, DetailParseResult):
                    det = cached
                else:
                    det = await parse_detail_page(detail_page, link)
                    if det.blocked:
                        detail_cache[link] = _DETAIL_BLOCKED
                    else:
                        detail_cache[link] = det

                if isinstance(detail_cache.get(link), DetailParseResult):
                    det = detail_cache[link]
                    if det.address:
                        address = det.address
                    evidence_extra = det.evidence or ""
                    row_district = resolve_district(
                        address, desc, title, evidence_extra, district
                    )

            if not price:
                continue

            records.append(
                {
                    "РАЙОН": row_district,
                    "КОМНАТЫ": parse_rooms(title),
                    "ПЛОЩАДЬ": parse_area(full_text),
                    "ЭТАЖ": parse_floor(full_text),
                    "АДРЕС": flatten_address_cell(address),
                    "ЦЕНА": price,
                    "ССЫЛКА": link,
                }
            )
        except Exception:
            continue

    return records


async def scrape_district(browser, district: str, base_url: str) -> list[dict]:
    ua = random.choice(USER_AGENTS)
    context = await browser.new_context(
        user_agent=ua,
        viewport={"width": random.randint(1280, 1920), "height": random.randint(768, 1080)},
        locale="ru-RU",
        timezone_id="Europe/Moscow",
        extra_http_headers={
            "Accept-Language": "ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Referer": "https://www.avito.ru/saransk",
        },
    )

    await context.add_init_script(STEALTH_JS)
    page = await context.new_page()
    detail_page = await context.new_page()
    detail_cache: dict = {}

    all_records = []
    consecutive_empty = 0
    prev_page_links = None

    print(f"\n{'=' * 50}")
    print(f"Район (выдача Avito): {district}")
    print(f"{'=' * 50}")

    try:
        await page.goto("https://www.avito.ru/saransk", wait_until="domcontentloaded", timeout=30000)
        await asyncio.sleep(random.uniform(2, 4))
    except Exception:
        pass

    for page_num in range(1, MAX_PAGES + 1):
        url = base_url if page_num == 1 else f"{base_url}{'&' if '?' in base_url else '?'}p={page_num}"
        print(f"  Страница {page_num}: ", end="", flush=True)

        try:
            await page.goto(url, wait_until="domcontentloaded", timeout=30000)
            await asyncio.sleep(random.uniform(1.5, 3.0))

            page_text = await page.inner_text("body")
            if page_looks_blocked(page_text):
                print("⚠️  Блокировка/IP/капча на выдаче. Пауза 60 с...")
                await asyncio.sleep(60)
                await page.goto(url, wait_until="domcontentloaded", timeout=30000)
                await asyncio.sleep(3)

            records = await parse_listing_page(
                page, district, detail_page=detail_page, detail_cache=detail_cache
            )
            print(f"{len(records)} объявлений", end="")

            if not records:
                consecutive_empty += 1
                prev_page_links = None
                if consecutive_empty >= 2:
                    print("  → Нет объявлений 2 страницы подряд, завершаем район")
                    break
            else:
                consecutive_empty = 0
                all_records.extend(records)
                links_now = frozenset(r.get("ССЫЛКА", "") for r in records)
                if page_num > 1 and prev_page_links is not None and links_now == prev_page_links:
                    print("  → Та же выдача — конец пагинации")
                    break
                prev_page_links = links_now

            if page_num == MAX_PAGES:
                print("  → Достигнут лимит MAX_PAGES")
                break

            delay = random.uniform(*PAGE_DELAY)
            print(f"  ⏱  пауза {delay:.1f}с")
            await asyncio.sleep(delay)

        except PwTimeout:
            print("  ⏰ Таймаут, пропускаем страницу")
            consecutive_empty += 1
            if consecutive_empty >= 3:
                break
        except Exception as e:
            print(f"  ❌ Ошибка: {e}")
            consecutive_empty += 1
            if consecutive_empty >= 3:
                break

    print(f"\n  ✅ Итого по выдаче «{district}»: {len(all_records)} объявлений")
    await context.close()
    return all_records


COLUMNS = ["РАЙОН", "КОМНАТЫ", "ПЛОЩАДЬ", "ЭТАЖ", "АДРЕС", "ЦЕНА", "ССЫЛКА"]
COL_WIDTHS = [18, 10, 10, 10, 48, 16, 60]


def save_xlsx(records: list[dict], path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Квартиры Саранск"
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    HEADER_FILL = PatternFill("solid", start_color="1F3864")
    HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    CELL_FONT = Font(name="Arial", size=10)
    CENTER = Alignment(horizontal="center", vertical="center")
    LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
    THIN = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    ROW_COLORS = ["FFFFFF", "EBF3FB"]

    ws.append(COLUMNS)
    for ci, (col, w) in enumerate(zip(COLUMNS, COL_WIDTHS), 1):
        c = ws.cell(1, ci)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = THIN
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 22

    for ri, rec in enumerate(records, 2):
        fill = PatternFill("solid", start_color=ROW_COLORS[ri % 2])
        for ci, col in enumerate(COLUMNS, 1):
            val = rec.get(col, "")
            if col in ("ЦЕНА",) and val:
                try:
                    val = int(val)
                except ValueError:
                    pass
            if col == "ПЛОЩАДЬ" and val:
                try:
                    val = float(val)
                except ValueError:
                    pass
            c = ws.cell(ri, ci, val)
            c.font = CELL_FONT
            c.fill = fill
            c.border = THIN
            c.alignment = CENTER if col in ("РАЙОН", "КОМНАТЫ", "ПЛОЩАДЬ", "ЭТАЖ", "ЦЕНА") else LEFT
        ws.row_dimensions[ri].height = 18

    wb.save(path)
    print(f"\n💾 Файл сохранён: {path}")
    print(f"📊 Всего строк: {len(records)}")


def record_quality(rec: dict) -> tuple:
    
    #Для дедупа: выше — строка с явным районом в адресе и полным адресом.
    addr = norm_lower(rec.get("АДРЕС") or "")
    dist = norm_spaces(rec.get("РАЙОН") or "")
    score = 0
    if extract_district_hint(addr):
        score += 4
    elif infer_microdistrict_from_street(addr):
        score += 3
    elif looks_like_street_address(addr):
        score += 2
    if len(addr) > 40:
        score += 1
    if dist and not is_forbidden_district_label(dist):
        score += 1
    return (score, len(addr))


async def main():
    all_records = []

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(
            headless=False,
            args=[
                "--no-sandbox",
                "--disable-blink-features=AutomationControlled",
                "--disable-infobars",
                "--window-size=1366,768",
            ],
        )

        for search in SEARCHES:
            records = await scrape_district(browser, search["district"], search["url"])
            all_records.extend(records)

            if search != SEARCHES[-1]:
                pause = random.uniform(8, 15)
                print(f"\n⏱  Пауза между районами: {pause:.0f}с")
                await asyncio.sleep(pause)

        await browser.close()

    by_link: dict[str, dict] = {}
    for r in all_records:
        key = r.get("ССЫЛКА", "") or ""
        if not key:
            continue
        prev = by_link.get(key)
        if prev is None or record_quality(r) > record_quality(prev):
            by_link[key] = r

    unique = list(by_link.values())
    unique.sort(key=lambda x: (x.get("РАЙОН") or "", x.get("АДРЕС") or ""))

    print(f"\n📋 Уникальных объявлений: {len(unique)} (дубликатов удалено: {len(all_records) - len(unique)})")
    save_xlsx(unique, OUTPUT_FILE)


if __name__ == "__main__":
    asyncio.run(main())
